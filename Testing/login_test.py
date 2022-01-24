from types import ClassMethodDescriptorType
from selenium import webdriver;
from selenium.webdriver.common.keys import Keys; 
import time
import xlsxwriter
import XLUtils
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.remote.webelement import WebElement
import unittest
from random import randint
from selenium.common.exceptions import NoSuchElementException
driver = webdriver.Chrome(executable_path='C:\selenium driver\chromedriver.exe')

driver.get("https://dsce-connect-ruddy.vercel.app/")
driver.maximize_window()
time.sleep(2)

a = ActionChains(driver)

s = driver.find_element_by_xpath("//a[normalize-space()='LOGIN']")
a.move_to_element(s).perform() 
time.sleep(0.3)

submit = driver.find_element_by_xpath("//a[normalize-space()='LOGIN']")
submit.click()

path = "C:\\Users\Solomon\Desktop\Login_Test.xlsx"
wb=openpyxl.load_workbook("C:\\Users\Solomon\Desktop\Login_Test.xlsx")
sh1 = wb.active

rows=XLUtils.getRowCount(path,'Sheet1')

def slow_type(element: WebElement, text: str, delay: float=0.2):
    """Send a text to an element one character at a time with a delay."""
    for character in text:
        element.send_keys(character)
        time.sleep(delay)

i=2
for r in range(2,rows+1):
    username=XLUtils.readData(path,'Sheet1',r,1)
    password=XLUtils.readData(path,'Sheet1',r,2)
    username1 = driver.find_element_by_xpath("/html/body/form/div[1]/input")        
    text = username
    slow_type(username1, text)
    password1 = driver.find_element_by_xpath("/html/body/form/div[2]/input")
    text = password
    slow_type(password1, text)
    submit = driver.find_element_by_xpath("/html/body/form/button")
    submit.click()

    if "http://localhost:3036/dashboard" == driver.current_url:
        sh1.cell(row = i,column=3,value="Test"+str(i)+"----->Passed :)")
        time.sleep(1)
        driver.back()
        driver.refresh()
    else:
        sh1.cell(row = i,column=3,value="Test"+str(i)+"----->Failed :(")
        time.sleep(1)
        driver.refresh()
    i=i+1
    time.sleep(0.3)
    driver.refresh()

wb.save("C:\\Users\Solomon\Desktop\Login_Test.xlsx")
time.sleep(0.3)
driver.quit()
